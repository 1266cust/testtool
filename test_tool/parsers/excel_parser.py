from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook


def read_excel_file(path: Path) -> str:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: List[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            clean = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if clean:
                parts.append(" | ".join(clean))
    return "\n".join(parts)